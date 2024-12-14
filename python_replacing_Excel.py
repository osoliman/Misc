import pandas as pd
from google.cloud import bigquery
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Initialize BigQuery client
client = bigquery.Client()

# First, get the new values from BigQuery
unique_identifiers = ['A', 'B']  # Get these from your actual data
query = f"""
SELECT 
    identifier,
    new_value
FROM your_project.your_dataset.your_table
WHERE identifier IN ({','.join([f"'{id}'" for id in unique_identifiers])})
"""

# Execute query and convert to DataFrame
bq_df = client.query(query).to_dataframe()
bq_df.columns = ['identifier', 'new_value']

# Create a dictionary for quick lookup
replacement_dict = dict(zip(bq_df['identifier'], bq_df['new_value']))

# Load the workbook and select the active sheet
wb = load_workbook('input.xlsx', data_only=False)
ws = wb.active

# Find column indices based on header names
header_row = 1  # Assuming headers are in first row
id_col_name = 'Race'
value_col_name = 'N_members'

# Find the column letters for our target columns
id_col_letter = None
value_col_letter = None

for col in range(1, ws.max_column + 1):
    col_letter = get_column_letter(col)
    cell_value = ws[f'{col_letter}{header_row}'].value
    if cell_value == id_col_name:
        id_col_letter = col_letter
    elif cell_value == value_col_name:
        value_col_letter = col_letter

if not id_col_letter or not value_col_letter:
    raise ValueError(f"Could not find columns: {id_col_name} and/or {value_col_name}")

# Update only the specific cells while preserving everything else
for row in range(2, ws.max_row + 1):  # Starting from row 2 (after header)
    identifier = ws[f'{id_col_letter}{row}'].value
    if identifier in replacement_dict:
        ws[f'{value_col_letter}{row}'].value = replacement_dict[identifier]

# Save the workbook
wb.save('output.xlsx')

# Print confirmation
print(f"Updated values in column {value_col_name} based on matches in column {id_col_name}")
